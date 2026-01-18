"""
Data Export Module

Provides functionality to export trade history, performance metrics,
and chart data in various formats (CSV, JSON, Excel).

Features:
- Export trade history
- Export performance metrics
- Export chart data (OHLCV)
- Multiple format support
- Date range filtering
- Compression support
- Email/download options

Usage:
    from src.web.data_export import DataExporter

    exporter = DataExporter()
    csv_data = exporter.export_trades(trades, format='csv')
    exporter.save_to_file(csv_data, 'trades.csv')
"""

import logging
import csv
import json
import io
import zipfile
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from pathlib import Path

logger = logging.getLogger(__name__)


class DataExporter:
    """Export trading data in various formats."""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize data exporter.

        Args:
            config: Exporter configuration
        """
        self.config = config or {}
        self.export_dir = Path(self.config.get("export_dir", "exports"))
        self.export_dir.mkdir(exist_ok=True)

        logger.info(f"Data exporter initialized (export dir: {self.export_dir})")

    def export_trades(
        self,
        trades: List[Dict[str, Any]],
        format: str = 'csv',
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> str:
        """
        Export trade history.

        Args:
            trades: List of trades
            format: Export format (csv, json, excel)
            start_date: Filter start date
            end_date: Filter end date

        Returns:
            Exported data as string
        """
        # Filter by date range
        filtered_trades = self._filter_by_date(trades, start_date, end_date)

        if format == 'csv':
            return self._export_trades_csv(filtered_trades)
        elif format == 'json':
            return self._export_trades_json(filtered_trades)
        elif format == 'excel':
            return self._export_trades_excel(filtered_trades)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def export_performance_metrics(
        self,
        stats: Dict[str, Any],
        format: str = 'json'
    ) -> str:
        """
        Export performance metrics.

        Args:
            stats: Performance statistics
            format: Export format

        Returns:
            Exported data as string
        """
        if format == 'json':
            return json.dumps(stats, indent=2)
        elif format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Metric', 'Value'])
            for key, value in stats.items():
                writer.writerow([key, value])
            return output.getvalue()
        else:
            raise ValueError(f"Unsupported format: {format}")

    def export_chart_data(
        self,
        chart_data: List[Dict[str, Any]],
        format: str = 'csv'
    ) -> str:
        """
        Export chart data (OHLCV or profit over time).

        Args:
            chart_data: Chart data points
            format: Export format

        Returns:
            Exported data as string
        """
        if format == 'csv':
            if not chart_data:
                return ""

            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=chart_data[0].keys())
            writer.writeheader()
            writer.writerows(chart_data)
            return output.getvalue()

        elif format == 'json':
            return json.dumps(chart_data, indent=2)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def export_bot_summary(
        self,
        bots: List[Dict[str, Any]],
        format: str = 'csv'
    ) -> str:
        """
        Export multi-bot summary.

        Args:
            bots: List of bot information
            format: Export format

        Returns:
            Exported data as string
        """
        if format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Bot ID', 'Strategy', 'Provider', 'Status',
                'Total Trades', 'Win Rate', 'Total Profit',
                'Uptime (seconds)', 'Created At'
            ])

            for bot in bots:
                writer.writerow([
                    bot.get('bot_id', ''),
                    bot.get('strategy', ''),
                    bot.get('provider', ''),
                    bot.get('status', ''),
                    bot.get('stats', {}).get('total_trades', 0),
                    f"{bot.get('stats', {}).get('win_rate', 0):.2f}%",
                    f"${bot.get('stats', {}).get('total_profit', 0):.2f}",
                    bot.get('uptime', 0),
                    bot.get('created_at', '')
                ])

            return output.getvalue()

        elif format == 'json':
            return json.dumps(bots, indent=2)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def save_to_file(
        self,
        data: str,
        filename: str,
        compress: bool = False
    ) -> Path:
        """
        Save exported data to file.

        Args:
            data: Data to save
            filename: Output filename
            compress: Whether to compress (zip)

        Returns:
            Path to saved file
        """
        filepath = self.export_dir / filename

        if compress:
            # Create zip file
            zip_filepath = filepath.with_suffix(filepath.suffix + '.zip')
            with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(filename, data)
            logger.info(f"Data exported to {zip_filepath} (compressed)")
            return zip_filepath
        else:
            # Save as plain file
            filepath.write_text(data)
            logger.info(f"Data exported to {filepath}")
            return filepath

    def _export_trades_csv(self, trades: List[Dict[str, Any]]) -> str:
        """Export trades to CSV format."""
        if not trades:
            return ""

        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            'Timestamp', 'Bot ID', 'Strategy', 'Pair', 'Side',
            'Size', 'Price', 'Profit', 'Fee', 'Exchange'
        ])

        # Rows
        for trade in trades:
            writer.writerow([
                trade.get('timestamp', ''),
                trade.get('bot_id', ''),
                trade.get('strategy', ''),
                trade.get('pair', ''),
                trade.get('side', ''),
                trade.get('size', 0),
                trade.get('price', 0),
                trade.get('profit', 0),
                trade.get('fee', 0),
                trade.get('exchange', '')
            ])

        return output.getvalue()

    def _export_trades_json(self, trades: List[Dict[str, Any]]) -> str:
        """Export trades to JSON format."""
        return json.dumps({
            'trades': trades,
            'count': len(trades),
            'exported_at': datetime.now().isoformat()
        }, indent=2)

    def _export_trades_excel(self, trades: List[Dict[str, Any]]) -> bytes:
        """Export trades to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trades"

            # Header
            headers = [
                'Timestamp', 'Bot ID', 'Strategy', 'Pair', 'Side',
                'Size', 'Price', 'Profit', 'Fee', 'Exchange'
            ]
            ws.append(headers)

            # Style header
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Data rows
            for trade in trades:
                ws.append([
                    trade.get('timestamp', ''),
                    trade.get('bot_id', ''),
                    trade.get('strategy', ''),
                    trade.get('pair', ''),
                    trade.get('side', ''),
                    trade.get('size', 0),
                    trade.get('price', 0),
                    trade.get('profit', 0),
                    trade.get('fee', 0),
                    trade.get('exchange', '')
                ])

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self._export_trades_csv(trades).encode('utf-8')

    def _filter_by_date(
        self,
        trades: List[Dict[str, Any]],
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> List[Dict[str, Any]]:
        """Filter trades by date range."""
        if not start_date and not end_date:
            return trades

        filtered = []
        for trade in trades:
            timestamp_str = trade.get('timestamp', '')
            if not timestamp_str:
                continue

            try:
                timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))

                if start_date and timestamp < start_date:
                    continue
                if end_date and timestamp > end_date:
                    continue

                filtered.append(trade)

            except (ValueError, AttributeError):
                logger.warning(f"Invalid timestamp: {timestamp_str}")
                continue

        return filtered

    def generate_profit_chart_data(
        self,
        trades: List[Dict[str, Any]],
        interval: str = '1h'
    ) -> List[Dict[str, Any]]:
        """
        Generate chart data for profit over time.

        Args:
            trades: List of trades
            interval: Time interval (1m, 5m, 15m, 1h, 4h, 1d)

        Returns:
            List of data points with timestamp and cumulative profit
        """
        if not trades:
            return []

        # Sort trades by timestamp
        sorted_trades = sorted(trades, key=lambda t: t.get('timestamp', ''))

        # Calculate interval in seconds
        interval_seconds = self._parse_interval(interval)

        # Group trades by interval
        chart_data = []
        current_profit = 0.0
        current_bucket_start = None

        for trade in sorted_trades:
            timestamp_str = trade.get('timestamp', '')
            if not timestamp_str:
                continue

            try:
                timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))

                # Initialize first bucket
                if current_bucket_start is None:
                    current_bucket_start = self._round_to_interval(timestamp, interval_seconds)

                # Check if we need a new bucket
                while timestamp >= current_bucket_start + timedelta(seconds=interval_seconds):
                    # Save current bucket
                    chart_data.append({
                        'timestamp': current_bucket_start.isoformat(),
                        'profit': current_profit
                    })
                    # Move to next bucket
                    current_bucket_start += timedelta(seconds=interval_seconds)

                # Add profit to current cumulative
                current_profit += trade.get('profit', 0)

            except (ValueError, AttributeError) as e:
                logger.warning(f"Invalid timestamp: {timestamp_str} - {e}")
                continue

        # Add final bucket
        if current_bucket_start:
            chart_data.append({
                'timestamp': current_bucket_start.isoformat(),
                'profit': current_profit
            })

        return chart_data

    def _parse_interval(self, interval: str) -> int:
        """Parse interval string to seconds."""
        unit = interval[-1]
        value = int(interval[:-1])

        if unit == 'm':
            return value * 60
        elif unit == 'h':
            return value * 3600
        elif unit == 'd':
            return value * 86400
        else:
            raise ValueError(f"Invalid interval: {interval}")

    def _round_to_interval(self, dt: datetime, interval_seconds: int) -> datetime:
        """Round datetime to interval."""
        timestamp = int(dt.timestamp())
        rounded = (timestamp // interval_seconds) * interval_seconds
        return datetime.fromtimestamp(rounded)


def create_data_exporter(config: Optional[Dict[str, Any]] = None) -> DataExporter:
    """
    Create and return data exporter instance.

    Args:
        config: Exporter configuration

    Returns:
        DataExporter instance
    """
    return DataExporter(config)
